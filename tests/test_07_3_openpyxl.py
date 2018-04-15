from pathlib import Path
from examples.c07_3_crud_sheets.openpyxl import index
from examples.c07_1_create_empty_excel_file.openpyxl \
    import index as create_index
import os
from openpyxl import Workbook
from base_test_cases import ExcelTest


class TestOpenPyXLCrud(ExcelTest):

    def test_list_all_sheets(self):
        # python35 cannot write to Path, so use os.path
        tests_path = os.path.dirname(
            os.path.abspath(__file__))
        project_root = os.path.dirname(
            tests_path)
        examples_path = os.path.join(project_root, 'examples')
        sample_xlsx_path = os.path.join(
            examples_path, 'c07_3_crud_sheets', '3_sheets.xlsx')
        self.assertTrue(
            Path(sample_xlsx_path).is_file())
        sheets = index.list_all_sheet_names(sample_xlsx_path)
        expected_sheets = ['Sheet1', 'Sheet2', 'Sheet3']
        self.assertListEqual(sheets, expected_sheets)

    def test_has_sheet(self):
        # python35 cannot write to Path, so use os.path
        tests_path = os.path.dirname(
            os.path.abspath(__file__))
        project_root = os.path.dirname(
            tests_path)
        examples_path = os.path.join(project_root, 'examples')
        sample_xlsx_path = os.path.join(
            examples_path, 'c07_3_crud_sheets', '3_sheets.xlsx')
        self.assertTrue(
            Path(sample_xlsx_path).is_file())
        self.assertTrue(index.has_sheet(sample_xlsx_path, 'Sheet1'))
        self.assertFalse(index.has_sheet(sample_xlsx_path, 'Does not exist'))

    def test_add_new_sheet(self):
        # python35 cannot write to Path, so use os.path
        save_path = os.path.join(
            os.path.dirname(
                os.path.abspath(__file__)),
            'c07_3_openpyxl_empty.xlsx')
        self.assertFalse(
            Path(save_path).is_file())
        create_index.save_empty_file(save_path)
        self.assertTrue(
            Path(save_path).is_file())
        new_sheet_name = 'New Sheet'
        self.assertFalse(index.has_sheet(save_path, new_sheet_name))
        index.add_new_sheet(save_path, new_sheet_name)
        self.assertTrue(index.has_sheet(save_path, new_sheet_name))
        # when attempt to add new sheet name clashes with existing
        # exception is thrown
        with self.assertRaises(Exception):
            index.add_new_sheet(save_path, new_sheet_name)

    def test_delete_sheet(self):
        # python35 cannot write to Path, so use os.path
        save_path = os.path.join(
            os.path.dirname(
                os.path.abspath(__file__)),
            'c07_3_openpyxl_empty.xlsx')
        self.assertFalse(
            Path(save_path).is_file())
        create_index.save_empty_file(save_path)
        new_sheet_name = 'New Sheet'
        index.add_new_sheet(save_path, new_sheet_name)
        self.assertTrue(index.has_sheet(save_path, new_sheet_name))
        index.delete_sheet(save_path, new_sheet_name)
        self.assertFalse(index.has_sheet(save_path, new_sheet_name))
        # when attempt to delete sheet name that does not exist
        # exception is thrown
        with self.assertRaises(Exception):
            index.delete_sheet(save_path, new_sheet_name)

    def test_order_sheets(self):
        # python35 cannot write to Path, so use os.path
        tests_path = os.path.dirname(
            os.path.abspath(__file__))
        project_root = os.path.dirname(
            tests_path)
        examples_path = os.path.join(project_root, 'examples')
        sample_xlsx_path = os.path.join(
            examples_path, 'c07_3_crud_sheets', '3_sheets_acb.xlsx')
        self.assertTrue(
            Path(sample_xlsx_path).is_file())
        save_path_abc = os.path.join(
            os.path.dirname(
                os.path.abspath(__file__)),
            'c07_3_3_sheets_abc.xlsx')
        self.assertFalse(
            Path(save_path_abc).is_file())
        index.order_sheets_alphabetically(
            sample_xlsx_path, dest_file_path=save_path_abc)
        self.assertTrue(
            Path(save_path_abc).is_file())

        sheets = index.list_all_sheet_names(sample_xlsx_path)
        expected_sheets = ['a', 'c', 'b']
        self.assertListEqual(sheets, expected_sheets)

        sheets = index.list_all_sheet_names(save_path_abc)
        expected_sheets = ['a', 'b', 'c']
        self.assertListEqual(sheets, expected_sheets)

        save_path_cba = os.path.join(
            os.path.dirname(
                os.path.abspath(__file__)),
            'c07_3_3_sheets_cba.xlsx')
        self.assertFalse(
            Path(save_path_cba).is_file())
        index.order_sheets_alphabetically(
            sample_xlsx_path,
            dest_file_path=save_path_cba,
            reverse=True
        )
        self.assertTrue(
            Path(save_path_cba).is_file())

        sheets = index.list_all_sheet_names(save_path_cba)
        expected_sheets = ['c', 'b', 'a']
        self.assertListEqual(sheets, expected_sheets)

    def test_rename_sheet(self):
        # python35 cannot write to Path, so use os.path
        tests_path = os.path.dirname(
            os.path.abspath(__file__))
        project_root = os.path.dirname(
            tests_path)
        examples_path = os.path.join(project_root, 'examples')
        sample_xlsx_path = os.path.join(
            examples_path, 'c07_3_crud_sheets', '3_sheets_acb.xlsx')
        self.assertTrue(
            Path(sample_xlsx_path).is_file())
        save_path_fcb = os.path.join(
            os.path.dirname(
                os.path.abspath(__file__)),
            'c07_3_3_sheets_fcb.xlsx')
        self.assertFalse(
            Path(save_path_fcb).is_file())
        index.rename_sheet(
            sample_xlsx_path,
            dest_file_path=save_path_fcb,
            sheet_name='a',
            new_sheet_name='f')
        self.assertTrue(
            Path(save_path_fcb).is_file())

        sheets = index.list_all_sheet_names(sample_xlsx_path)
        expected_sheets = ['a', 'c', 'b']
        self.assertListEqual(sheets, expected_sheets)

        sheets = index.list_all_sheet_names(save_path_fcb)
        expected_sheets = ['f', 'c', 'b']
        self.assertListEqual(sheets, expected_sheets)

    def test_read_sheet_data(self):
        # python35 cannot write to Path, so use os.path
        tests_path = os.path.dirname(
            os.path.abspath(__file__))
        project_root = os.path.dirname(
            tests_path)
        examples_path = os.path.join(project_root, 'examples')
        sample_xlsx_path = os.path.join(
            examples_path, 'c07_3_crud_sheets', '3_sheets_acb.xlsx')
        self.assertTrue(
            Path(sample_xlsx_path).is_file())

        sheet_name = 'a'

        data = index.read_sheet_data(sample_xlsx_path, sheet_name)
        cols, rows = 1, 1
        expected_data = [[0 for x in range(cols)] for y in range(rows)]
        expected_data[0][0] = 'a'
        self.assertListEqual(data, expected_data)

        non_existent_sheet = 'non_existent_sheet'
        with self.assertRaises(Exception):
            data = index.read_sheet_data(sample_xlsx_path, non_existent_sheet)

    def test_write_sheet_data(self):
        # python35 cannot write to Path, so use os.path
        cols, rows = 1, 1
        expected_data = [[0 for x in range(cols)] for y in range(rows)]
        expected_data[0][0] = 'a'
        workbook = Workbook()
        sheetname = 'a'
        after_written_workbook = index.write_sheet_data(
            workbook, sheetname, expected_data
        )
        saved_sheet = after_written_workbook[sheetname]
        savedcols, savedrows = saved_sheet.max_column, saved_sheet.max_row
        saved_data = [[0 for x in range(savedcols)] for y in range(savedrows)]
        for row_index, row in enumerate(
            saved_sheet.iter_rows(
                max_col=saved_sheet.max_column,
                max_row=saved_sheet.max_row
            )
        ):
            for col_index, cell in enumerate(row):
                saved_data[row_index][col_index] = cell.value
        self.assertListEqual(expected_data, saved_data)

        with self.assertRaises(Exception):
            after_written_workbook = index.write_sheet_data(
                after_written_workbook, sheetname, expected_data
            )
