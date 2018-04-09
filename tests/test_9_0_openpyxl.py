from openpyxl.styles import colors, Font
from examples.c90_font_styles.openpyxl import index
from openpyxl import Workbook
from base_test_cases import ExcelTest


class TestOpenPyXLFontStyles(ExcelTest):

    def test_font_color(self):
        wb = Workbook()
        ws = wb.active
        a1 = ws['A1']
        a1_font = a1.font
        default_color = colors.Color(
            indexed=None, type='theme', rgb=None, tint=0.0, theme=1, auto=None)
        self.assertEqual(a1_font.color, default_color)
        a1.font = Font(color="FF000000")
        self.assertEqual(a1.font.color.rgb, "FF000000")
        a1 = index.set_font_color_red(wb)
        self.assertEqual(a1.font.color.rgb, colors.RED)

    def test_font_size(self):
        wb = Workbook()
        ws = wb.active
        a1 = ws['A1']
        a1_font = a1.font
        default_size = 11
        self.assertEqual(a1_font.size, default_size)
        new_size = 20
        a1 = index.set_font_size(wb, new_size)
        self.assertEqual(a1.font.size, new_size)

    def test_font_style(self):
        wb = Workbook()
        ws = wb.active
        a1 = ws['A1']
        a1_font = a1.font
        default_bold = False
        default_italic = False
        default_underline = None
        default_style = 'Calibri'
        self.assertEqual(a1_font.bold, default_bold)
        self.assertEqual(a1_font.italic, default_italic)
        self.assertEqual(a1_font.underline, default_underline)
        self.assertEqual(a1_font.name, default_style)
        new_style = "Helvetica"
        a1 = index.set_font_style(wb, new_style)
        self.assertEqual(a1.font.name, new_style)
        self.assertTrue(a1.font.bold)
        self.assertTrue(a1.font.italic)
        four_kinds_of_underlines = [
            'single', 'singleAccounting',
            'double',  'doubleAccounting']
        self.assertEqual(a1.font.underline, four_kinds_of_underlines[0])
