from openpyxl.styles import PatternFill
from examples.c09_2_fill_color.openpyxl import index
from openpyxl import Workbook
from base_test_cases import ExcelTest


class TestOpenPyXLFillColor(ExcelTest):

    def test_font_color(self):
        wb = Workbook()
        ws = wb.active
        a1 = ws['A1']
        a1_fill = a1.fill
        default_fill = PatternFill()
        self.assertEqual(a1_fill, default_fill)
        a1 = index.set_fill_color_green(wb)
        expected_fill = PatternFill(
            fill_type=None,
            start_color='FFFFFFFF',
            end_color='FF000000')
        self.assertEqual(a1.fill, expected_fill)
