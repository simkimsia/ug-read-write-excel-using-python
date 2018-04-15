from examples.c10_1_set_print_area.openpyxl import index
from examples.c07_1_create_empty_excel_file.openpyxl \
    import index as create_index
from openpyxl import load_workbook
from pathlib import Path
import os
from base_test_cases import ExcelTest


class TestOpenPyXLPrintArea(ExcelTest):

    def test_set_print_area(self):
        tests_path = os.path.dirname(
            os.path.abspath(__file__))
        write_xlsx_path = os.path.join(
            tests_path, 'set_print_area.xlsx')
        create_index.save_empty_file(write_xlsx_path)
        self.assertTrue(
            Path(write_xlsx_path).is_file())

        workbook = load_workbook(write_xlsx_path)
        worksheet = workbook.active
        self.assertIsNone(worksheet.print_area)
        index.set_print_area(write_xlsx_path, 'A1:U50')
        wb_after_set_print_area = load_workbook(write_xlsx_path)
        ws_after_set_print_area = wb_after_set_print_area.active
        expected_result = ['$A$1:$U$50']
        self.assertEqual(ws_after_set_print_area.print_area, expected_result)
