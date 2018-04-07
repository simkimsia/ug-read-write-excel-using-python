from pathlib import Path
from examples.c71_create_empty_excel_file.openpyxl \
    import index as create_index
from examples.c91_charts.openpyxl import index
from openpyxl import load_workbook
from base_test_cases import ExcelTest
import os


class TestOpenPyXLFontStyles(ExcelTest):

    def test_barcharts(self):
        # python35 cannot write to Path, so use os.path
        tests_path = os.path.dirname(
            os.path.abspath(__file__))
        write_xlsx_path = os.path.join(
            tests_path, 'make_charts.xlsx')
        create_index.save_empty_file(write_xlsx_path)
        self.assertTrue(
            Path(write_xlsx_path).is_file())
        wb = load_workbook(write_xlsx_path)
        ws = wb.active
        self.assertTrue(len(ws._charts) == 0)
        index.make_bar_chart(write_xlsx_path)
        wb = load_workbook(write_xlsx_path)
        ws = wb.active
        self.assertTrue(len(ws._charts) > 0)
